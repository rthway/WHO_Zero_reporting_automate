from openpyxl import load_workbook
import mysql.connector
import csv 


wb = load_workbook(filename='who.xlsx')
sheet = wb.worksheets[0]

mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="password!",
  database="openmrs"
)

mycursor = mydb.cursor()

sql = "SELECT count(cn.name) as Disease_Count, \
DATE_SUB(CURDATE(),INTERVAL 7 DAY) as start_date, \
DATE_SUB(CURDATE(),INTERVAL 1 DAY) as end_date, \
CURDATE() as date, \
week(CURDATE())as week_no \
FROM obs o \
  INNER JOIN concept_name cn \
    ON cn.concept_id = o.value_coded AND cn.name = 'Pneumonia' AND cn.concept_name_type = 'FULLY_SPECIFIED' AND \
       cn.voided IS FALSE and date(o.obs_datetime) between DATE_SUB(CURDATE(),INTERVAL 7 DAY) AND DATE_SUB(CURDATE(),INTERVAL 1 DAY);"

mycursor.execute(sql)

myresult = mycursor.fetchall()


for x in myresult:    
    sheet.cell(row=42,column=1).value=x[0]
    sheet.cell(row=13,column=7).value=x[1]
    sheet.cell(row=13,column=9).value=x[2]
    sheet.cell(row=45,column=7).value=x[3]
    sheet.cell(row=13,column=2).value=x[4]

wb.save('who.xlsx') 